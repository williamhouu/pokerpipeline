"""Layer 6: Explanation Generator.

The only layer that calls an external LLM. Every other layer is deterministic
Python; this one takes a fully populated SpotData (Layer 5) and asks the model
to *narrate* the solver-derived facts in the team's coaching voice.

Per the brief's "the LLM never thinks about poker" principle, the model is not
allowed to reason strategically here: it receives the decision class, concept
tags, equity numbers, range shape, and the correct action, and writes prose
for the six question-facing CSV columns:

    option_1, option_2, option_3, option_4, correct_answer, answer_explanation

Voice is locked by an 8-rule system prompt distilled from the team's gold
examples (docs/output_format_examples.xlsx Sheet 2), an in-context exemplar
block of ~8 of those gold rows, and the brief's banned-phrase list (no em
dashes, no semicolons, no corporate phrasing). Output is JSON for parsing.

Option style is chosen from solver signals before the call -- binary action
(Call/Fold), frequency (Always/Mostly + composite labels), or sizing
(33%/75%/135%) --
and passed to the LLM as an explicit instruction so the model doesn't pick.

Validation: the `correct_answer` string must equal one of `option_1`...
`option_4` exactly. A mismatch triggers one corrective retry; a second failure
raises `ExplanationValidationError` so Layer 7 can route the question to human
review.

Usage:

    from pipeline.explanation_generator import generate_explanation
    explanation = generate_explanation(spot_data)        # reads ANTHROPIC_API_KEY

The Anthropic SDK call uses prompt caching on the system block and the gold
example block, since both are identical across thousands of generations.
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from pipeline.fact_extractor.archetypes import (
    ARCHETYPE_GUIDANCE, RECOMMENDED_ACTION_ARCHETYPES,
)
from pipeline.fact_extractor.spot_data import SpotData

# Default Anthropic model. Opus 4.7 is the default (June 2026, the team's
# call) and the production model for every batch. Callers (admin panel,
# scripts) can override `model` with Sonnet 4.6 as the cheap/fast option for
# iterating on prompts before committing to a real batch.
DEFAULT_MODEL = "claude-opus-4-7"
DEFAULT_TEMPERATURE = 0.3                  # tight enough to stay on-voice
DEFAULT_MAX_TOKENS = 2000                  # 4 short options + a multi-paragraph explanation
GOLD_EXAMPLE_COUNT = 8                     # brief: "8-12 gold examples"

# Models that have removed/deprecated the `temperature` parameter. The
# Anthropic API returns 400 invalid_request_error if you include
# `temperature` when calling any of these. We drop the param at call time
# and let the model use its default sampling. Add new entries as
# deprecations land. (Opus 4.x also removes top_p/top_k -- we never send
# those.)
MODELS_WITHOUT_TEMPERATURE: frozenset[str] = frozenset({
    "claude-opus-4-8",
    "claude-opus-4-7",
    "claude-opus-4-5",
})


def call_messages_create(client, *, model, max_tokens, temperature, system, messages):
    """Wrapper around ``client.messages.create`` that handles per-model
    parameter compatibility.

    One quirk today: Opus 4.x removed ``temperature`` -- including it 400s
    the whole batch. We drop the param for models in
    :data:`MODELS_WITHOUT_TEMPERATURE` (passed through unchanged for Sonnet
    etc.).

    Lives here (not in batch.py or per-call site) so every Layer 6
    entry point -- postflop, preflop full, preflop answer-only, PLO --
    routes through one place. Adding a new compatibility quirk is a
    one-line update.
    """
    kwargs = {
        "model": model,
        "max_tokens": max_tokens,
        "system": system,
        "messages": messages,
    }
    if model not in MODELS_WITHOUT_TEMPERATURE:
        kwargs["temperature"] = temperature
    return client.messages.create(**kwargs)
GOLD_XLSX = (Path(__file__).resolve().parent.parent
             / "docs" / "output_format_examples.xlsx")

# Banned-phrase patterns -- the brief lists em dashes, semicolons, and "corporate
# phrasing." The corporate-phrasing list is seeded from common LLM tells; the
# Layer 7 voice checker will grow it from observed failures.
BANNED_LITERAL_PHRASES = (
    "—",                                   # em dash
    ";",                                   # semicolon
    "let's dive in", "let us dive in",
    "in conclusion",
    "it is important to note",
    "it's important to note",
    "at the end of the day",
    "leverage",                            # corporate verb
    "synergy",
    "moving forward",
    "circle back",
)


# --- the eight voice rules ---------------------------------------------------
# Distilled from the gold examples in docs/output_format_examples.xlsx Sheet 2,
# "Golden explanation examples - I". Embedded in the system prompt verbatim.
VOICE_RULES = [
    # 1. Verdict first
    "Open with the verdict. The first sentence states the correct action "
    "plainly (\"The best play is to call.\", \"You should always check here.\", "
    "\"This is a clear fold.\"). Do not bury the answer.",
    # 2. Second-person coaching
    "Address the reader directly in second person (\"you\", \"your hand\", "
    "\"you're holding\"). Never refer to \"the player\" or \"hero\" -- the "
    "student IS hero.",
    # 3. Structure: verdict -> reasoning -> optional exploit caveat
    "Structure the body as: (a) verdict, (b) the reasoning (range, equity, "
    "position, board), (c) an optional one-line exploitative note when the "
    "spot is population-sensitive. Skip (c) if it doesn't apply.",
    # 4. Length: 2-5 sentences
    "Keep the explanation between 2 and 5 sentences. Tight is better than "
    "thorough. If you can say it in 3 sentences, say it in 3.",
    # 5. Concrete naming
    "Name things concretely: positions by abbreviation (UTG, HJ, CO, BTN, "
    "SB, BB), hands by class (top pair top kicker, open-ended straight draw, "
    "third pair), board features (paired, monotone, broadway-heavy).",
    # 6. Plain-language solver facts
    "Translate solver numbers into plain language. Say \"you have the "
    "stronger range here\" rather than \"hero range advantage = 4.2%\"; say "
    "\"you have very little of your range that can continue\" rather than "
    "\"continuing range frequency 18%\".",
    # 7. No em dashes, no semicolons
    "Never use an em dash. Never use a semicolon. Use periods, commas, and "
    "short sentences. Do not use any phrase from the BANNED PHRASES list.",
    # 8. Confident instructive tone
    "Write with the confidence of a coach who has solved the spot. Do not "
    "hedge with \"it might be\", \"perhaps\", \"in theory\". State what the "
    "right play is and why, plainly.",
    # 9. Suit emojis for specific card references (Ryan-feedback Fix 3, May 2026)
    "When referencing a SPECIFIC card or two-card combo in prose, use the "
    "suit emoji form: A♠️, K❤️, T♦️, "
    "9♣️ (rank letter immediately followed by the suit emoji, no "
    "space). Never use plain-text suit letters like 'Kh', 'Ad', or 'JsTs' in "
    "the answer explanation -- those are internal solver notation, not voice. "
    "This matches the suit-emoji convention already used in the Question "
    "column. Hand-class names (\"top pair top kicker\", \"open-ended straight "
    "draw\") stay as plain prose -- the emoji rule only applies when you "
    "actually name specific cards.",
    # 10. Name specific villain combos (Ryan-feedback Fix 4, May 2026)
    "When the explanation discusses villain's range (what villain has, what "
    "villain is continuing with, what villain can show up with), name 2 to 3 "
    "SPECIFIC hand classes from `range_data.villain_top_value_combos` in the "
    "SOLVER DATA block, citing each one's `example_combos` directly. Use the "
    "emoji form already in the data (e.g. 'villain shows up with K♠️K♣️ and "
    "J♦️J♣️ for sets, plus A♠️K♠️ for the strongest two-pair'). Do not "
    "describe villain's range abstractly with only generic phrases like "
    "'value hands' or 'top of range' -- anchor the abstract phrase to actual "
    "combos from the field. The top entries in that field are villain's "
    "highest-leverage hands at this node, sorted by weight*strength.",
]


# --- Ryan-feedback Fix 3 supplementary gold examples (May 2026) --------------
# These pin the suit-emoji convention (rule 9) by showing concrete villain
# combo callouts in the team voice. The .xlsx gold pool predates the convention
# so we ship these inline; the cached prompt block carries them on every call.
SUPPLEMENTARY_EXAMPLES_SUIT_EMOJI = (
    "When BTN raises the flop and barrels turn, the strongest combos he can "
    "show up with are K♠️K♣️ and J♦️J♣️ "
    "for sets, plus K❤️J❤️ and Q❤️J❤️ "
    "as the strongest two-pair combos.",
    "BB's flop check-raise range here is loaded with two-pair: 8♠️"
    "5♦️, 8♦️3♦️, and 5♣️3♣️ "
    "are all in there at full weight.",
    "Villain's continuing range is mostly draws: 9❤️7❤️, "
    "T❤️8❤️, A♦️K♦️. None of those "
    "have you in bad shape on this river.",
)

# --- Ryan-feedback Fix 5 archetype-aware gold examples (May 2026) ------------
# Per the user spec, six new gold explanations covering the most-confused
# archetypes from the V7.1 audit. Each example is structured as
# (archetype, hand_stage, hand_class_label, prose) so we can render them
# grouped by archetype in the cached prompt block. The trap_check example
# is the canonical V7.1 failure-mode: strong hand on a later street where
# hero had the betting lead, recommended check, prior LLM frame was wrong.
SUPPLEMENTARY_EXAMPLES_ARCHETYPE = (
    {"archetype": "trap_check",
     "hand_stage": "Turn",
     "hand_class": "set_no_draws",
     "prose":
         "The best play is to check here. You raised the flop and got "
         "called, then spiked a set on the 8♠️ turn. Your hand is too "
         "strong to fold villain out by betting -- BB will fold the "
         "9❤️7❤️, T❤️9♣️ semi-bluffs you want to keep in. Letting BB "
         "barrel river turns those bluffs into chips. BB's continuing "
         "range here is mostly K♣️J♣️, T♠️T♣️ for sets, plus 6♥6♣ and "
         "the straight combos 4♦️3♦️, 7♣️6♣️ -- you cover all of them, "
         "and only the straights would call a turn bet."},
    {"archetype": "pot_control_check",
     "hand_stage": "Turn",
     "hand_class": "top_pair_good_kicker_no_draws",
     "prose":
         "Check it back. Top pair good kicker plays best in a controlled "
         "pot here: betting builds a pot you don't want with one pair, "
         "BB's worse hands like K♦️T♣️ and Q♠️J♠️ all fold, and BB's "
         "better hands (A♣️K♣️, K♠️K♦️) keep going. Checking takes "
         "the free showdown card with your second-pair-class equity."},
    {"archetype": "bluff_catch",
     "hand_stage": "River",
     "hand_class": "second_pair_no_draws",
     "prose":
         "Call. BB is overbetting a polarised range: A♠️A♣️ and "
         "K♦️K♥ for the value, against missed flush draws like J♣️9♣️ "
         "and 7♦️6♦️ for the bluffs. You only need 33% to call this "
         "price, and second pair beats every bluff combo BB is using. "
         "This is not a value call -- you're calling because the price "
         "is right vs the bluff frequency, not because your hand is "
         "strong."},
    {"archetype": "protection_bet",
     "hand_stage": "Flop",
     "hand_class": "top_pair_weak_kicker_no_draws",
     "prose":
         "Bet 33% of the pot. Top pair with a weak kicker needs to deny "
         "equity to BB's continuing range -- the gutshots Q♠️J♠️, "
         "J❤️T❤️ and the under-pair pairs 6♦️6♠️, 5♣️5♥ all "
         "realise meaningful equity if you check. A small bet folds out "
         "the air and charges the draws, without bloating the pot vs "
         "the slowplays."},
    {"archetype": "bluff",
     "hand_stage": "Flop",
     "hand_class": "no_pair_air_with_flush_draw",
     "prose":
         "Bet 75% pot. You have nine outs to the nut flush plus three "
         "overs, and BB's range here is mostly weak pairs and "
         "broadway-air that folds to a turn barrel. K♣️J♣️ and "
         "Q♣️J♣️ continue, but the bulk of villain's range (the "
         "Q♥️J♥, J♦️T♦️, 9♠️8♠️ class) folds today, and the "
         "draw equity is the backup when called."},
    {"archetype": "fold_to_polar",
     "hand_stage": "River",
     "hand_class": "two_pair_top_no_draws",
     "prose":
         "Fold. BB has overbet polar on a board that completed straights "
         "and flushes -- the value combos K❤️Q❤️, 7♦️6♦️, T♣️9♣️ "
         "all crush you, and BB's range is too thin on bluffs to justify "
         "calling at this price. Top two looks strong, but the overbet "
         "is sized exactly for value here."},
)


# --- public dataclass --------------------------------------------------------
@dataclass
class GeneratedExplanation:
    """The six LLM-written columns for one CSV row.

    `option_3` and `option_4` are empty strings when the decision has fewer
    than four options. `correct_answer` always equals one of `option_1`...
    `option_4` exactly (enforced by `_validate`).
    """

    option_1: str
    option_2: str
    option_3: str
    option_4: str
    correct_answer: str
    answer_explanation: str

    def options(self) -> list[str]:
        """The non-empty option strings, in order."""
        return [opt for opt in (self.option_1, self.option_2,
                                self.option_3, self.option_4) if opt]


class ExplanationValidationError(RuntimeError):
    """Raised when the LLM's output fails validation after one retry.

    Layer 7 catches this and routes the spot to human review.

    Optionally carries the LLM's most recent attempt so the batch
    layer can surface it to a human reviewer with full context:
    the raw text the model produced, plus the parsed
    :class:`GeneratedExplanation` if parsing succeeded but a downstream
    validator rejected it.

    Attributes:
        last_attempt_text: Raw text from the most recent LLM response.
            Empty string when no API call happened (rare; usually a
            pre-call error like a bad correct_answer).
        last_attempt_candidate: Parsed GeneratedExplanation from the
            most recent attempt, or None if parsing never succeeded.
    """

    def __init__(
        self,
        message: str,
        *,
        last_attempt_text: str = "",
        last_attempt_candidate: GeneratedExplanation | None = None,
    ) -> None:
        super().__init__(message)
        self.last_attempt_text = last_attempt_text
        self.last_attempt_candidate = last_attempt_candidate


# --- frequency-to-prefix mapping (deterministic, no LLM) ---------------------
# Brackets for the Always/Mostly prefix in frequency-style options. Inclusive
# at the lower bound: freq >= 0.95 -> "Always", freq >= 0.05 -> "Mostly",
# below 0.05 -> empty string (the action is essentially not played).
#
# Pre-Apr-2026 the mapping had four brackets (Always/Mostly/Sometimes/Rarely).
# Ryan-feedback item #2: a standalone "Sometimes X" option is ambiguous --
# two such options ("Sometimes check", "Sometimes bet") aren't mutually
# exclusive readings of the strategy. Collapsing the lower three brackets
# into a single "Mostly" makes the 2-action option set "Always A / Mostly A
# / Mostly B / Always B" universally; 3+ action spots use composite labels
# like "Mostly X, sometimes Y" instead of standalone "Sometimes" labels
# (see Fix 2b in the option-style instruction and validators).
#
# Used by:
#   * _expected_correct_prefix to compute the hard-constraint prefix the LLM
#     receives in the frequency-style option-style instruction;
#   * validators.validate_correct_answer_verb to check the LLM honoured it.
_FREQ_PREFIX_BRACKETS = (
    (0.95, "Always"),
    (0.05, "Mostly"),
)


def frequency_to_verb_prefix(freq: float) -> str:
    """Map a Pio frequency to its deterministic option-label prefix.

      [0.95, 1.0]  -> "Always"
      [0.05, 0.95) -> "Mostly"
      [0,    0.05) -> ""               (action essentially not played)

    "Sometimes" and "Rarely" used to be separate prefixes; collapsed into
    "Mostly" per Ryan's Apr-2026 V6 review (item #2). The LLM uses "sometimes"
    only as the secondary verb of a composite label like "Mostly call,
    sometimes raise" -- never as a standalone option prefix.
    """
    for floor, label in _FREQ_PREFIX_BRACKETS:
        if freq >= floor:
            return label
    return ""


# --- option-style detection (deterministic, no LLM) --------------------------
_BET_VERBS = ("bet", "raise", "donk", "overbet", "shove", "jam")


def _detect_option_style(spot_data: SpotData) -> str:
    """Pick the option style from solver signals.

    Returns one of:
      * "sizing"        -- multiple bet/raise sizes available (e.g. flop bet 33%
                           pot vs 75% pot). Options should list pot fractions
                           or chip amounts.
      * "frequency"     -- a single dominant action below 80% frequency: the
                           strategy is meaningfully mixed. Options should be
                           Always X / Mostly X / Mostly Y / Always Y.
      * "binary_action" -- a clearly-dominant single action with no sizing
                           choice. Options are the raw verbs (Call, Fold,
                           Raise) like the team's simpler gold rows.
    """
    bet_sizes = [frac for action, frac in
                 spot_data.decision_data.option_pot_fractions.items()
                 if any(v in action for v in _BET_VERBS) and frac > 0]
    if len({round(f, 2) for f in bet_sizes}) >= 2:
        return "sizing"
    strategy = spot_data.decision_data.range_aggregate_strategy
    if strategy and max(strategy.values()) < 0.80:
        return "frequency"
    return "binary_action"


def _expected_correct_prefix(spot_data: SpotData) -> str | None:
    """The deterministic prefix the LLM must use for correct_answer.

    Applies only to frequency-style spots -- binary_action and sizing styles
    use bare verbs or sizing labels with no prefix. Returns None when no
    prefix constraint applies.
    """
    if _detect_option_style(spot_data) != "frequency":
        return None
    strategy = spot_data.decision_data.range_aggregate_strategy
    if not strategy:
        return None
    return frequency_to_verb_prefix(max(strategy.values()))


def _top_two_verbs(spot_data: SpotData) -> tuple[str, str] | None:
    """The two highest-frequency verbs from Pio's range strategy, or None.

    Used by the frequency-style instruction to pin the option set's verb
    coverage (so the LLM cannot silently drop an action the solver plays --
    the Row 1 defect from the May-2026 audit).
    """
    strategy = spot_data.decision_data.range_aggregate_strategy
    if not strategy or len(strategy) < 2:
        return None
    ranked = sorted(strategy.items(), key=lambda kv: kv[1], reverse=True)
    return ranked[0][0], ranked[1][0]


def _option_style_instruction(style: str, spot_data: SpotData) -> str:
    """The natural-language instruction the LLM gets for option formatting."""
    if style == "sizing":
        fractions = sorted({round(f * 100)
                            for action, f in
                            spot_data.decision_data.option_pot_fractions.items()
                            if any(v in action for v in _BET_VERBS) and f > 0})
        examples = ", ".join(f"\"{p}% pot\"" for p in fractions[:4]) or \
            "\"33% pot\", \"75% pot\", \"135% pot\""
        return (
            "OPTION STYLE: sizing. The decision is between bet sizes. Each "
            f"option must be a sizing label such as {examples}. If the "
            "solver also offers a check or fold, include that as option_1; "
            "list the bet sizes in ascending order after it. Use 2 to 4 "
            "options total; leave unused options as empty strings."
        )
    if style == "frequency":
        prefix = _expected_correct_prefix(spot_data) or "Mostly"
        # Pio actions with non-trivial frequency. The 3+ action branch
        # (Ryan-feedback Fix 2b) kicks in when at least three verbs are
        # played at >= 5% -- a true multi-way mix that the 4-option
        # Always/Mostly template can't represent without standalone
        # "Sometimes X" labels (banned per Ryan's V6 review).
        strategy = spot_data.decision_data.range_aggregate_strategy
        meaningful = [v for v, f in
                      sorted(strategy.items(), key=lambda kv: -kv[1])
                      if f >= 0.05]
        if len(meaningful) >= 3:
            verb_a = meaningful[0]
            others = ", ".join(repr(v) for v in meaningful[1:])
            return (
                "OPTION STYLE: frequency (3+ actions). The solver mixes "
                "between three or more actions; the 4-option "
                "Always/Mostly template would force a standalone "
                "\"Sometimes X\" label, which is banned per the Apr-2026 "
                "review (those labels read as ambiguous to players). "
                f"Use COMPOSITE labels of the form \"Mostly {verb_a}, "
                f"sometimes Y\" pairing the dominant verb {verb_a!r} with "
                "each secondary verb. Templates:\n"
                f"  option_1 = \"Always {verb_a}\"\n"
                f"  option_2 = \"Mostly {verb_a}, sometimes <secondary>\"\n"
                f"  option_3 = \"Mostly {verb_a}, sometimes <other secondary>\"\n"
                f"  option_4 = \"Always <secondary>\" or empty\n"
                f"Pio plays {verb_a!r} as the dominant action and also "
                f"{others} at >= 5%. Cover each secondary action in at "
                f"least one composite label; do NOT emit a standalone "
                f"\"Sometimes X\" label.\n\n"
                f"HARD CONSTRAINT -- correct_answer must equal exactly "
                f"\"{prefix} {verb_a}\" if the dominant action's frequency "
                f"is below 95%, or \"Always {verb_a}\" if it is at or above "
                f"95%. The prefix {prefix!r} is computed deterministically "
                f"by Python from Pio's range frequency."
            )
        top_two = _top_two_verbs(spot_data)
        verb_coverage = ""
        if top_two:
            verb_a, verb_b = top_two
            verb_coverage = (
                f" The two highest-frequency verbs in Pio's range strategy "
                f"are {verb_a!r} (dominant) and {verb_b!r}. Your option set "
                f"MUST include both of these verbs across the four options "
                f"-- do not drop either one in favour of a more elegant "
                f"template (this is the Row-1 defect the May-2026 audit "
                f"caught: the LLM offered call-vs-raise options when the "
                f"actual strategy was call-vs-fold)."
            )
        return (
            "OPTION STYLE: frequency. The solver mixes between two actions. "
            "Use exactly four options in this template: \"Always <action A>\", "
            "\"Mostly <action A>\", \"Mostly <action B>\", \"Always "
            "<action B>\", where action A is the action the solver picks "
            "more often. Both middle options use \"Mostly\" -- standalone "
            "\"Sometimes X\" / \"Rarely X\" labels are banned (per Ryan's "
            "Apr-2026 review: ambiguous, not mutually exclusive).\n\n"
            f"HARD CONSTRAINT -- correct_answer must equal exactly "
            f"\"{prefix} <action A>\" where <action A> is the verb of Pio's "
            f"highest-frequency action. The prefix {prefix!r} is computed "
            f"deterministically by Python from Pio's range frequency; do "
            f"NOT substitute your own judgement (e.g. do not pick \"Always\" "
            f"when Python says \"Mostly\").{verb_coverage}"
        )
    return (
        "OPTION STYLE: binary action. The decision has a clear best action "
        "with no sizing nuance. Each option must be a single verb: \"Call\", "
        "\"Fold\", \"Check\", \"Bet\", \"Raise\". Use 2 or 3 options that match "
        "the actions the solver offers; leave unused options as empty strings."
    )


# --- prompt assembly ---------------------------------------------------------
def _format_voice_rules() -> str:
    return "\n".join(f"{i + 1}. {rule}" for i, rule in enumerate(VOICE_RULES))


def _format_banned_phrases() -> str:
    return ", ".join(f"\"{p}\"" for p in BANNED_LITERAL_PHRASES)


def _format_supplementary_examples() -> str:
    parts = [f"  EX{i + 1}. {ex}"
             for i, ex in enumerate(SUPPLEMENTARY_EXAMPLES_SUIT_EMOJI)]
    return "\n".join(parts)


def _format_archetype_examples() -> str:
    parts = []
    for i, ex in enumerate(SUPPLEMENTARY_EXAMPLES_ARCHETYPE, start=1):
        parts.append(
            f"  ARCH{i}. archetype={ex['archetype']}  "
            f"({ex['hand_stage']}, hand_class={ex['hand_class']})\n"
            f"    {ex['prose']}")
    return "\n".join(parts)


def _format_archetype_catalog() -> str:
    parts = []
    for archetype in RECOMMENDED_ACTION_ARCHETYPES:
        parts.append(f"  - {archetype}: {ARCHETYPE_GUIDANCE[archetype]}")
    return "\n".join(parts)


def build_system_prompt() -> str:
    """The static system prompt: voice rules, banned phrases, output schema.

    Stable across every call, so the SDK call marks it cacheable.
    """
    return (
        "You are the explanation generator for a poker training app. You "
        "translate solver-derived facts into short, on-voice coaching "
        "explanations. You never reason about poker yourself: every "
        "strategic claim in your output must be supported by a field in "
        "the SOLVER DATA block the user gives you.\n\n"
        "VOICE RULES (all nine apply to every output):\n"
        f"{_format_voice_rules()}\n\n"
        "SUIT-EMOJI CITATION EXAMPLES (Ryan-feedback Fix 3, May 2026 -- "
        "these supplement the .xlsx gold pool, which predates the emoji "
        "convention). When you cite specific combos in the answer "
        "explanation, match this voice and notation:\n"
        f"{_format_supplementary_examples()}\n\n"
        "STRATEGIC ARCHETYPES (Ryan-feedback Fix 5, May 2026). The data "
        "block carries a `decision_data.recommended_action_archetype` field. "
        "It is the STRATEGIC FRAME your explanation must be built around -- "
        "the action itself is in `decision_data.correct_action`. The 13 "
        "archetypes and the frame each one demands:\n"
        f"{_format_archetype_catalog()}\n\n"
        "ARCHETYPE-FRAMED GOLD EXAMPLES (Ryan-feedback Fix 5, May 2026 -- "
        "these target the archetype-mismatch failures from the V7.1 audit, "
        "especially trap_check on later streets where hero had the betting "
        "lead). Match the strategic frame each one carries:\n"
        f"{_format_archetype_examples()}\n\n"
        f"BANNED PHRASES (never appear in any output field): {_format_banned_phrases()}.\n\n"
        "DETERMINISTIC FREQUENCY PREFIX MAPPING: When a frequency-style "
        "option-style instruction specifies a required prefix for "
        "correct_answer, use it exactly. The prefix is computed in Python "
        "from Pio's range frequency of the dominant action:\n"
        "  >= 95%   -> \"Always\"\n"
        "  5-95%    -> \"Mostly\"\n"
        "  < 5%     -> the action is not played; do not emit a prefix.\n"
        "Per Apr-2026 review (Ryan): standalone \"Sometimes X\" and \"Rarely "
        "X\" labels are BANNED. They are ambiguous (two such options are not "
        "mutually exclusive readings of the strategy). In 3+ action spots "
        "use composite labels like \"Mostly call, sometimes raise\" instead "
        "-- the option-style instruction shows the template.\n"
        "Do NOT substitute your own judgement (e.g. do not write "
        "\"Always check\" when Python supplied \"Mostly\").\n\n"
        "AGGREGATE vs PER-COMBO: the SOLVER DATA field "
        "`range_mean_evs_per_action` is the RANGE-WEIGHTED mean EV across "
        "hero's whole range -- not the specific hero combo's EV. Never "
        "attribute these numbers to a specific combo (\"Ks3s specifically "
        "has EV +27bb\" is wrong; \"on average across your range, betting "
        "is worth +27bb\" is right).\n\n"
        "OUTPUT FORMAT: respond with a single JSON object and nothing else. "
        "No prose before or after, no markdown fences. The object has exactly "
        "these six keys, all string-valued:\n"
        "  option_1, option_2, option_3, option_4, correct_answer, answer_explanation\n"
        "Empty options must be empty strings (\"\"), not null. The "
        "correct_answer string must equal one of option_1..option_4 "
        "character-for-character; the consumer pairs them by exact string match."
    )


def _trim_spot_data(spot_data: SpotData) -> dict:
    """A compact view of the SpotData for the user prompt.

    The full to_dict() is heavy with empty Combo lists; the LLM only needs the
    decision-relevant fields. Empty collections and zero floats are dropped to
    keep the prompt readable.

    The 169-entry ip_range_snapshot / oop_range_snapshot dicts (Ryan ask, May
    2026) are dropped entirely -- they're for Layer 8's UI-range-grid CSV
    columns, not for the LLM. Leaving them in would add ~3KB of mostly-zero
    weights to every prompt for no decision-relevant signal (Layer 6 already
    gets `villain_top_value_combos` for the actionable per-class view).
    """
    raw = spot_data.to_dict()
    # Drop the two snapshot fields before the recursive trim so they cannot
    # leak into the LLM data block.
    raw.pop("ip_range_snapshot", None)
    raw.pop("oop_range_snapshot", None)

    def trim(value):
        if isinstance(value, dict):
            return {k: trim(v) for k, v in value.items()
                    if v not in ("", None, [], {}, 0.0, 0)}
        if isinstance(value, list):
            return [trim(item) for item in value]
        return value

    # Always keep `street` even if "preflop" -- it isn't optional.
    trimmed = trim(raw)
    trimmed.setdefault("spot_metadata", {})["street"] = raw["spot_metadata"]["street"]
    return trimmed


def _question_framing(spot_data: SpotData) -> str:
    """The plain-English description of what's being decided at this spot."""
    meta = spot_data.spot_metadata
    decision = spot_data.decision_data
    hero = meta.hero_position or "hero"
    villain = meta.villain_position or "villain"
    actions = ", ".join(decision.options) or "(no options recorded)"
    framing = (
        f"Stage: {meta.street}. Hero ({hero}) is deciding against "
        f"{villain}. Available actions in the solver: {actions}. The "
        f"solver-correct action is \"{decision.correct_action}\" "
        f"(frequency dominant). The explanation must justify exactly "
        f"that action."
    )
    # Ryan-feedback Fix 5 (May 2026): when the classifier picked an archetype,
    # surface its specific guidance in the live prompt so the LLM can't miss
    # the strategic frame in the catalog. The trap_check guidance in
    # particular has explicit anti-pattern instructions ("DO NOT frame as
    # 'pot control' / 'villain has the nut advantage'") that need to be
    # this close to the data block for V7.1's failure mode to stop recurring.
    archetype = decision.recommended_action_archetype
    if archetype and archetype in ARCHETYPE_GUIDANCE:
        framing += (
            f"\n\nRECOMMENDED-ACTION ARCHETYPE: {archetype}.\n"
            f"{ARCHETYPE_GUIDANCE[archetype]}"
        )
    return framing


@lru_cache(maxsize=1)
def load_gold_examples(path: str | None = None) -> tuple[dict, ...]:
    """Load the gold question/explanation pairs from the .xlsx sample.

    Cached so repeated generations don't re-parse the file. Returns a tuple of
    plain dicts with the in-prompt fields the model needs to mimic; non-data
    cells are skipped. The path argument is the cache key (a string so the
    decorator can hash it).
    """
    import openpyxl                            # imported lazily -- demo-only dep

    src = Path(path) if path else GOLD_XLSX
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Golden explanation examples - I"]
    headers = [c.value for c in ws[1]]
    keep = ("Hand Stage", "Question", "option 1", "option 2", "option 3",
            "option 4", "Correct Answer", "Answer Explanation",
            "Preflop Pot Type", "Pot Participant", "Stack Depth")
    examples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if not (record.get("Question") and record.get("Answer Explanation")):
            continue
        examples.append({k: (record.get(k) or "") for k in keep})
    return tuple(examples)


def _format_gold_examples(examples) -> str:
    """Render the gold-example block for the user prompt."""
    parts = []
    for index, ex in enumerate(examples, start=1):
        options = [ex["option 1"], ex["option 2"], ex["option 3"], ex["option 4"]]
        parts.append(
            f"--- GOLD EXAMPLE {index} ({ex['Hand Stage'] or 'unspecified'}) ---\n"
            f"Question: {ex['Question']}\n"
            f"Options: {options}\n"
            f"Correct Answer: {ex['Correct Answer']}\n"
            f"Answer Explanation:\n{ex['Answer Explanation']}"
        )
    return "\n\n".join(parts)


def build_user_prompt(spot_data: SpotData,
                      gold_examples,
                      style: str) -> str:
    """The per-call user prompt: gold examples + framing + solver data.

    Returned as one string so the SDK call can split it into a cached
    gold-examples block and a per-call live block.
    """
    return (
        "Read these gold examples to lock in the voice. Imitate the cadence, "
        "verdict-first structure, and concrete naming. Do not copy phrasing.\n\n"
        f"{_format_gold_examples(gold_examples)}\n\n"
        "=== NEW QUESTION TO WRITE ===\n\n"
        f"{_question_framing(spot_data)}\n\n"
        f"{_option_style_instruction(style, spot_data)}\n\n"
        "SOLVER DATA (every strategic claim in your output must trace back "
        "to a field below; do not invent equity numbers or range claims):\n"
        f"{json.dumps(_trim_spot_data(spot_data), indent=2, default=str)}\n\n"
        "Now write the JSON object."
    )


def _build_messages_payload(spot_data: SpotData,
                            gold_examples,
                            style: str) -> tuple[list, list]:
    """The (system, messages) payload for client.messages.create.

    Splits the user prompt into a cached gold-example block and a live block
    so the Anthropic prompt cache hits across thousands of generations.
    """
    system = [{
        "type": "text",
        "text": build_system_prompt(),
        "cache_control": {"type": "ephemeral"},
    }]
    cached_block = (
        "Read these gold examples to lock in the voice. Imitate the cadence, "
        "verdict-first structure, and concrete naming. Do not copy phrasing.\n\n"
        f"{_format_gold_examples(gold_examples)}"
    )
    live_block = (
        "\n\n=== NEW QUESTION TO WRITE ===\n\n"
        f"{_question_framing(spot_data)}\n\n"
        f"{_option_style_instruction(style, spot_data)}\n\n"
        "SOLVER DATA (every strategic claim in your output must trace back "
        "to a field below; do not invent equity numbers or range claims):\n"
        f"{json.dumps(_trim_spot_data(spot_data), indent=2, default=str)}\n\n"
        "Now write the JSON object."
    )
    messages = [{
        "role": "user",
        "content": [
            {"type": "text", "text": cached_block,
             "cache_control": {"type": "ephemeral"}},
            {"type": "text", "text": live_block},
        ],
    }]
    return system, messages


# --- response parsing & validation -------------------------------------------
_JSON_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)


def parse_response(text: str) -> GeneratedExplanation:
    """Parse the LLM's response (a JSON object) into a GeneratedExplanation.

    Tolerates accidental code fences (the system prompt forbids them but
    Sonnet still emits them occasionally) and a leading prose preamble: we
    scan for the first { ... } object and parse that.
    """
    cleaned = _JSON_FENCE_RE.sub("", text).strip()
    # Pull out the first JSON object if there's extra prose.
    if not cleaned.startswith("{"):
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start < 0 or end <= start:
            raise ExplanationValidationError(
                f"no JSON object in LLM response: {text!r}")
        cleaned = cleaned[start:end + 1]
    try:
        # strict=False: accept raw control characters (literal newlines)
        # inside string values -- models writing multi-line explanations
        # sometimes emit them unescaped, and the newline IS the wanted prose.
        data = json.loads(cleaned, strict=False)
    except json.JSONDecodeError as exc:
        raise ExplanationValidationError(
            f"LLM response was not valid JSON: {exc}") from exc

    required = ("option_1", "option_2", "option_3", "option_4",
                "correct_answer", "answer_explanation")
    missing = [k for k in required if k not in data]
    if missing:
        raise ExplanationValidationError(
            f"LLM response missing keys: {missing}")
    return GeneratedExplanation(**{k: (data[k] or "") for k in required})


def _validate(explanation: GeneratedExplanation) -> str | None:
    """Return None if valid, else a short error message for the retry prompt."""
    options = explanation.options()
    if not options:
        return "the response had no non-empty options"
    if not explanation.correct_answer:
        return "correct_answer was empty"
    if explanation.correct_answer not in options:
        return (f"correct_answer {explanation.correct_answer!r} did not "
                f"match any of the four options exactly: {options!r}")
    if not explanation.answer_explanation.strip():
        return "answer_explanation was empty"
    return None


# --- main entry point --------------------------------------------------------
def generate_explanation(spot_data: SpotData, *,
                         client=None,
                         model: str = DEFAULT_MODEL,
                         temperature: float = DEFAULT_TEMPERATURE,
                         max_tokens: int = DEFAULT_MAX_TOKENS,
                         gold_examples=None,
                         max_retries: int = 1) -> GeneratedExplanation:
    """Run Layer 6 on a SpotData and return the six LLM-written columns.

    Reads ``ANTHROPIC_API_KEY`` from the environment when constructing the
    default client. Pass a mock ``client`` (any object with ``messages.create``)
    for tests. ``gold_examples`` defaults to the team's .xlsx gold pool; pass a
    list for tests to avoid hitting the file.

    Retries once with a corrective message when the response fails validation
    (most commonly: ``correct_answer`` not matching any ``option_N`` exactly).
    A second failure raises ``ExplanationValidationError`` so Layer 7 can flag
    the question for human review.
    """
    if client is None:
        from anthropic import Anthropic        # imported lazily -- runtime dep
        client = Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    if gold_examples is None:
        gold_examples = list(load_gold_examples())[:GOLD_EXAMPLE_COUNT]

    style = _detect_option_style(spot_data)
    system, messages = _build_messages_payload(spot_data, gold_examples, style)

    # Layer 7 validators (pipeline.validators) need the DecisionData; we run
    # them in the retry loop AFTER the structural _validate check, so a
    # malformed JSON fails before we ever look at the strategic content.
    from pipeline.validators import run_audit_validators

    last_error: str | None = None
    for attempt in range(max_retries + 1):
        response = call_messages_create(
            client,
            model=model, max_tokens=max_tokens, temperature=temperature,
            system=system, messages=messages,
        )
        text = _extract_text(response)
        try:
            explanation = parse_response(text)
            error = _validate(explanation)
            if error is None:
                # Structural check passed; now the Layer 7 strategic checks.
                audit_result = run_audit_validators(explanation,
                                                    spot_data.decision_data)
                if audit_result.is_valid:
                    return explanation
                error = audit_result.error_message
            last_error = error
        except ExplanationValidationError as exc:
            last_error = str(exc)

        # One retry with a corrective assistant + user turn. Append, don't
        # restart -- keeps the gold-example cache block warm.
        messages = messages + [
            {"role": "assistant", "content": text},
            {"role": "user", "content":
             "That response failed validation: "
             f"{last_error}. Re-emit the JSON object, correcting only that "
             "issue. Keep the same options and explanation otherwise."},
        ]

    raise ExplanationValidationError(
        f"Layer 6 failed after {max_retries + 1} attempts; last error: "
        f"{last_error}. Spot routed to human review.")


def _extract_text(response) -> str:
    """Pull the assistant text out of an Anthropic Messages response.

    Tolerates both the SDK's content-block list and a raw string (handy for
    mocking in tests).
    """
    if isinstance(response, str):
        return response
    content = getattr(response, "content", None)
    if content is None:
        raise ExplanationValidationError(
            f"Anthropic response had no content: {response!r}")
    # SDK content is a list of TextBlock objects; we want the first text block.
    for block in content:
        text = getattr(block, "text", None)
        if text is not None:
            return text
        if isinstance(block, dict) and "text" in block:
            return block["text"]
    raise ExplanationValidationError(
        f"no text block in Anthropic response: {content!r}")


# --- helper for the demo / tests: turn an explanation into row overrides -----
def explanation_to_row_overrides(explanation: GeneratedExplanation) -> dict:
    """The six column values Layer 8's CSV row should pick up from Layer 6.

    Format Writer fills these in place of its `[TBD by Layer 6]` placeholders.
    """
    return {
        "option 1": explanation.option_1,
        "option 2": explanation.option_2,
        "option 3": explanation.option_3,
        "option 4": explanation.option_4,
        "Correct Answer": explanation.correct_answer,
        "Answer Explanation": explanation.answer_explanation,
    }


__all__ = [
    "GeneratedExplanation",
    "ExplanationValidationError",
    "VOICE_RULES",
    "BANNED_LITERAL_PHRASES",
    "DEFAULT_MODEL",
    "GOLD_XLSX",
    "build_system_prompt",
    "build_user_prompt",
    "parse_response",
    "generate_explanation",
    "explanation_to_row_overrides",
    "load_gold_examples",
    "frequency_to_verb_prefix",
]
